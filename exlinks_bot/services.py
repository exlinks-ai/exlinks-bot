from __future__ import annotations

import csv
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from .config import PLANS


NAME_HEADERS = {"name", "product_name", "title", "ad", "mehsul", "məhsul"}
LINK_HEADERS = {"link", "url", "product_link", "amazon_link", "kechid", "keçid"}


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def format_dt(dt: datetime | None) -> str:
    if dt is None:
        return "-"

    # SQLite bəzən timezone məlumatını itirib naive datetime qaytarır.
    # Bu layihədə bütün tarixlər UTC yaradıldığı üçün naive gəlirsə UTC kimi qəbul edirik.
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

def next_delivery_for_plan(plan_code: str, base: datetime | None = None) -> datetime:
    base = base or utc_now()
    plan = PLANS[plan_code]
    days = int(plan["days"])
    return base + timedelta(days=days)


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _extract_name_link(row: dict[str, object]) -> dict[str, str]:
    name = ""
    link = ""

    for key, value in row.items():
        if key in NAME_HEADERS and not name:
            name = str(value or "").strip()
        if key in LINK_HEADERS and not link:
            link = str(value or "").strip()

    return {"name": name, "link": link}


def _rows_from_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(2048)
        fh.seek(0)

        try:
            has_header = csv.Sniffer().has_header(sample)
        except csv.Error:
            has_header = True

        rows: list[dict[str, str]] = []

        if has_header:
            reader = csv.DictReader(fh)
            for row in reader:
                normalized = {_normalize_header(k): str(v or "").strip() for k, v in row.items()}
                rows.append(_extract_name_link(normalized))
            return rows

        reader = csv.reader(fh)
        for row in reader:
            if len(row) < 2:
                continue
            rows.append({"name": str(row[0]).strip(), "link": str(row[1]).strip()})
        return rows


def _rows_from_excel(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    first_row = next(rows_iter, None)

    if first_row is None:
        return []

    headers = [_normalize_header(cell) for cell in first_row]
    has_name = any(header in NAME_HEADERS for header in headers)
    has_link = any(header in LINK_HEADERS for header in headers)

    rows: list[dict[str, str]] = []

    if has_name and has_link:
        for raw_row in rows_iter:
            row_map = {headers[idx]: raw_row[idx] for idx in range(min(len(headers), len(raw_row)))}
            rows.append(_extract_name_link(row_map))
        return rows

    first_values = [str(value or "").strip() for value in first_row]
    if len(first_values) >= 2:
        rows.append({"name": first_values[0], "link": first_values[1]})

    for raw_row in rows_iter:
        values = [str(value or "").strip() for value in raw_row]
        if len(values) >= 2:
            rows.append({"name": values[0], "link": values[1]})

    return rows


def parse_product_file(path: str | Path) -> list[dict[str, str]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return _rows_from_csv(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_excel(file_path)

    raise ValueError("Unsupported file type. Use .xlsx, .xlsm or .csv")


def build_admin_user_tag(username: str | None, first_name: str | None, telegram_id: int) -> str:
    display = (first_name or username or f"User {telegram_id}").strip()
    if username:
        return f"{display} (@{username})"
    return f"{display} (ID: {telegram_id})"